import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Paleta de Colores"

# ── helpers ────────────────────────────────────────────────────────────────────
def hex_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def thin_border():
    s = Side(style="thin", color="1E2330")
    return Border(left=s, right=s, top=s, bottom=s)

def cell(ws, row, col, value, bold=False, font_color="D1D4DC", bg=None, align="left", size=11):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", bold=bold, color=font_color, size=size)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = thin_border()
    if bg:
        c.fill = hex_fill(bg)
    return c

# ── data ───────────────────────────────────────────────────────────────────────
sections = [
    {
        "title": "FONDOS",
        "label_color": "787B86",
        "rows": [
            ("Fondo principal",  "#0A0E17", "0A0E17", "D1D4DC", "Background del app, inputs, selectbox"),
            ("Fondo tarjetas",   "#131722", "131722", "D1D4DC", "Cards, sidebar, charts, alertas, info boxes"),
            ("Fondo hover",      "#1A1F2E", "1A1F2E", "D1D4DC", "Hover state de metric cards"),
            ("Bordes / divisores","#1E2330","1E2330", "D1D4DC", "Todos los bordes, gridlines de gráficas, HR"),
        ]
    },
    {
        "title": "ACENTOS / ESTADO  ← candidatos para colores de marca",
        "label_color": "F5A623",
        "rows": [
            ("Acento principal (amber)", "#F5A623", "F5A623", "0A0E17",
             "Valores de métricas, botones, borde top de cards, series de gráficas — EL MÁS VISIBLE"),
            ("Verde (positivo)",        "#00D4AA", "00D4AA", "0A0E17",
             "Botón de descarga, indicadores de éxito, series de gráficas"),
            ("Rojo (alerta/error)",     "#FF4757", "FF4757", "FFFFFF",
             "Alertas, errores, flags/bugs, series de gráficas"),
            ("Azul (neutro/info)",      "#3B82F6", "3B82F6", "FFFFFF",
             "Headers de sección, series de gráficas"),
        ]
    },
    {
        "title": "TEXTO",
        "label_color": "787B86",
        "rows": [
            ("Texto principal",  "#D1D4DC", "D1D4DC", "0A0E17", "Cuerpo, contenido principal, valores"),
            ("Texto secundario", "#787B86", "787B86", "0A0E17", "Labels, ejes de gráficas, descripciones"),
            ("Texto apagado",    "#4A4E5A", "4A4E5A", "D1D4DC", "Notas técnicas, texto terciario"),
        ]
    },
]

# ── layout ─────────────────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 28   # nombre
ws.column_dimensions["B"].width = 14   # hex actual
ws.column_dimensions["C"].width = 14   # muestra de color
ws.column_dimensions["D"].width = 22   # hex de marca (vacío para el diseñador)
ws.column_dimensions["E"].width = 14   # muestra marca (vacío)
ws.column_dimensions["F"].width = 52   # dónde se usa

# title row
ws.row_dimensions[1].height = 36
for col in range(1, 7):
    c = ws.cell(row=1, column=col)
    c.fill = hex_fill("131722")
    c.border = thin_border()

title_cell = ws.cell(row=1, column=1, value="INVENTARIO DE COLORES — ALY DASHBOARD")
title_cell.font = Font(name="Calibri", bold=True, color="F5A623", size=14)
title_cell.alignment = Alignment(horizontal="left", vertical="center")
ws.merge_cells("A1:F1")

# header row
ws.row_dimensions[2].height = 22
headers = ["Nombre", "Hex actual", "Muestra actual", "Hex de marca\n(rellenar)", "Muestra\nmarca", "Dónde se usa"]
for i, h in enumerate(headers, 1):
    cell(ws, 2, i, h, bold=True, font_color="0A0E17", bg="F5A623", align="center", size=10)

# sections
current_row = 3
for section in sections:
    # section header
    ws.row_dimensions[current_row].height = 20
    sc = ws.cell(row=current_row, column=1, value=section["title"])
    sc.font = Font(name="Calibri", bold=True, color=section["label_color"], size=10)
    sc.fill = hex_fill("0A0E17")
    sc.alignment = Alignment(horizontal="left", vertical="center")
    sc.border = thin_border()
    ws.merge_cells(f"A{current_row}:F{current_row}")
    current_row += 1

    for nombre, hex_str, fill_hex, text_on_swatch, uso in section["rows"]:
        ws.row_dimensions[current_row].height = 36

        # col A: nombre
        cell(ws, current_row, 1, nombre, bold=False, font_color="D1D4DC", bg="131722")

        # col B: hex value
        cell(ws, current_row, 2, hex_str, bold=True, font_color="F5A623", bg="131722", align="center")

        # col C: color swatch
        c = ws.cell(row=current_row, column=3)
        c.fill = hex_fill(fill_hex)
        c.value = hex_str
        c.font = Font(name="Calibri", color=text_on_swatch, bold=True, size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()

        # col D: hex de marca — vacío para el diseñador
        cell(ws, current_row, 4, "", bg="1A1F2E", align="center")

        # col E: muestra marca — vacío
        c2 = ws.cell(row=current_row, column=5)
        c2.fill = hex_fill("0A0E17")
        c2.border = thin_border()

        # col F: uso
        cell(ws, current_row, 6, uso, font_color="787B86", bg="131722")

        current_row += 1

# instructions row
ws.row_dimensions[current_row + 1].height = 28
note = ws.cell(row=current_row + 1, column=1,
               value="→ El diseñador debe completar la columna D (Hex de marca) y la columna E quedará como referencia visual.")
note.font = Font(name="Calibri", italic=True, color="787B86", size=10)
note.fill = hex_fill("0A0E17")
note.alignment = Alignment(horizontal="left", vertical="center")
note.border = thin_border()
ws.merge_cells(f"A{current_row+1}:F{current_row+1}")

output = "/Users/daniel/Desktop/colores_aly_dashboard.xlsx"
wb.save(output)
print(f"Guardado en: {output}")
